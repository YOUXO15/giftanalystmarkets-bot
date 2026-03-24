"""Local file export business logic."""

from __future__ import annotations

import asyncio
import csv
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO, StringIO
import re
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from src.config.settings import Settings
from src.db.repositories.deal_repo import DealRepository
from src.db.repositories.export_log_repo import ExportLogRepository
from src.db.repositories.settings_repo import SettingsRepository
from src.db.repositories.ton_rate_repo import TonRateRepository
from src.db.repositories.user_repo import UserRepository
from src.integrations.ton_client import TonClient
from src.utils.currency_conversion import convert_amount, normalize_currency
from src.utils.enums import Currency, ExportFormat
from src.utils.helpers import build_registration_required_text

DEFAULT_EXPORT_FIELDS = [
    "id",
    "item",
    "gift_number",
    "market",
    "buy",
    "sell",
    "fee",
    "profit",
    "margin",
    "currency",
    "ton_rate",
    "status",
    "opened_at",
    "closed_at",
    "created_at",
    "updated_at",
    "gift_link",
]

ALLOWED_FIELD_KEYS = set(DEFAULT_EXPORT_FIELDS) | {"category"}
ALLOWED_STATUS_KEYS = {"open", "closed", "cancelled"}
ALLOWED_PROFIT_FILTERS = {"any", "positive", "negative", "zero", "non_negative", "non_positive"}
ALLOWED_QUERY_KEYS = {"format", "status", "currency", "profit", "days", "fields", "limit"}

_EXPORT_LOCKS: defaultdict[int, asyncio.Lock] = defaultdict(asyncio.Lock)


@dataclass(slots=True)
class ExportFileResult:
    """Represents a generated export file."""

    success: bool
    message: str
    filename: str | None = None
    content: bytes | None = None
    mime_type: str | None = None


@dataclass(slots=True)
class ExportQuery:
    """User-defined parameters for custom export."""

    statuses: set[str] | None = None
    currencies: set[str] | None = None
    report_currency: str | None = None
    profit_filter: str = "any"
    days: int | None = None
    fields: list[str] | None = None
    limit: int | None = None


def build_export_query_help_text() -> str:
    """Return a user-facing guide for parameterized export."""

    return (
        "<b>Экспорт по параметрам</b>\n"
        "Отправь параметры в формате key=value через `;`.\n\n"
        "Доступные ключи:\n"
        "format=csv|xlsx\n"
        "status=open,closed,cancelled\n"
        "currency=USD,EUR,RUB,TON,USDT\n"
        "profit=any|positive|negative|zero|non_negative|non_positive\n"
        "days=30\n"
        "limit=100\n"
        "fields=id,item,gift_number,market,buy,sell,fee,profit,margin,currency,ton_rate,status,opened_at,closed_at,created_at,updated_at,gift_link\n\n"
        "Пример:\n"
        "<code>format=xlsx;status=closed;currency=USD,TON;profit=positive;days=60;fields=id,item,gift_number,market,buy,sell,profit,margin,status</code>\n\n"
        "Для отмены отправь: <code>отмена</code>"
    )


def parse_export_query_text(raw_text: str) -> tuple[ExportQuery | None, ExportFormat | None, str | None]:
    """Parse user-provided export query string into typed filters."""

    text = raw_text.strip()
    if not text:
        return None, None, "Пустой запрос параметров. Отправь параметры в формате key=value."

    chunks = [part.strip() for part in re.split(r"[;\n]+", text) if part.strip()]
    if not chunks:
        return None, None, "Параметры не распознаны. Используй формат key=value."

    query = ExportQuery()
    export_format = ExportFormat.CSV

    for chunk in chunks:
        if "=" not in chunk:
            return None, None, f"Неверный параметр `{chunk}`. Используй формат key=value."

        raw_key, raw_value = chunk.split("=", 1)
        key = raw_key.strip().lower()
        value = raw_value.strip()
        if not key or not value:
            return None, None, f"Неверный параметр `{chunk}`. Ключ и значение обязательны."
        if key not in ALLOWED_QUERY_KEYS:
            supported = ", ".join(sorted(ALLOWED_QUERY_KEYS))
            return None, None, f"Неизвестный ключ `{key}`. Доступные: {supported}."

        if key == "format":
            try:
                export_format = ExportFormat(value.lower())
            except ValueError:
                return None, None, "Параметр format должен быть `csv` или `xlsx`."
            continue

        if key == "status":
            values = {item.strip().lower() for item in value.split(",") if item.strip()}
            if not values:
                return None, None, "Параметр status не может быть пустым."
            if "all" in values:
                query.statuses = None
            elif not values.issubset(ALLOWED_STATUS_KEYS):
                allowed = ", ".join(sorted(ALLOWED_STATUS_KEYS))
                return None, None, f"Неверный status. Используй: {allowed}."
            else:
                query.statuses = values
            continue

        if key == "currency":
            values = {item.strip().upper() for item in value.split(",") if item.strip()}
            allowed_currencies = {member.value for member in Currency}
            if not values:
                return None, None, "Параметр currency не может быть пустым."
            if "ALL" in values:
                query.currencies = None
            elif not values.issubset(allowed_currencies):
                allowed = ", ".join(sorted(allowed_currencies))
                return None, None, f"Неверная валюта. Используй: {allowed}."
            else:
                query.currencies = values
            continue

        if key == "profit":
            profit_filter = value.lower()
            if profit_filter not in ALLOWED_PROFIT_FILTERS:
                allowed = ", ".join(sorted(ALLOWED_PROFIT_FILTERS))
                return None, None, f"Неверный profit-фильтр. Используй: {allowed}."
            query.profit_filter = profit_filter
            continue

        if key == "days":
            try:
                days = int(value)
            except ValueError:
                return None, None, "Параметр days должен быть целым числом."
            if days <= 0 or days > 3650:
                return None, None, "Параметр days должен быть в диапазоне 1..3650."
            query.days = days
            continue

        if key == "limit":
            try:
                limit = int(value)
            except ValueError:
                return None, None, "Параметр limit должен быть целым числом."
            if limit <= 0 or limit > 5000:
                return None, None, "Параметр limit должен быть в диапазоне 1..5000."
            query.limit = limit
            continue

        if key == "fields":
            values: list[str] = []
            for item in value.split(","):
                field_key = item.strip().lower()
                if not field_key:
                    continue
                if field_key not in ALLOWED_FIELD_KEYS:
                    allowed = ",".join(DEFAULT_EXPORT_FIELDS)
                    return None, None, f"Неизвестное поле `{field_key}`. Доступные: {allowed}."
                if field_key not in values:
                    values.append(field_key)
            if not values:
                return None, None, "Параметр fields не может быть пустым."
            query.fields = values

    return query, export_format, None


class ExportService:
    """Application service for local CSV/XLSX exports."""

    def __init__(
        self,
        session_maker: async_sessionmaker[AsyncSession],
        settings: Settings | None = None,
    ) -> None:
        self._session_maker = session_maker
        self._settings = settings
        self._ton_client = TonClient(settings) if settings is not None else None

    async def export_deals_file(
        self,
        telegram_id: int,
        export_format: ExportFormat,
        query: ExportQuery | None = None,
    ) -> ExportFileResult:
        """Prepare deal rows and build a local export file."""

        active_query = query or ExportQuery()
        lock = _EXPORT_LOCKS[telegram_id]

        async with lock:
            async with self._session_maker() as session:
                user_repo = UserRepository(session)
                settings_repo = SettingsRepository(session)
                deal_repo = DealRepository(session)
                export_log_repo = ExportLogRepository(session)

                user = await user_repo.get_by_telegram_id(telegram_id)
                if user is None:
                    return ExportFileResult(success=False, message=build_registration_required_text())

                user_settings = await settings_repo.get_by_user_id(user.id)
                if user_settings is None:
                    await settings_repo.create_default(user.id)
                    await session.commit()

                deals = await deal_repo.get_all_by_user(user.id)
                if not deals:
                    return ExportFileResult(
                        success=False,
                        message="Нет данных для экспорта. Сначала добавь подарок через кнопку «Добавить подарок».",
                    )

                filtered_deals = self._apply_query_filters(deals, active_query)
                if not filtered_deals:
                    return ExportFileResult(
                        success=False,
                        message="По выбранным параметрам сделки не найдены.",
                    )

                if self._settings is not None:
                    export_count_today = await self._get_today_export_count(export_log_repo, user.id)
                    if export_count_today >= self._settings.daily_export_limit:
                        return ExportFileResult(
                            success=False,
                            message=self._build_daily_limit_message(export_count_today),
                        )

                ton_rate = await self._get_ton_rate_for_export()
                rows = self._build_export_rows(
                    filtered_deals,
                    ton_rate,
                    selected_fields=active_query.fields,
                    report_currency=active_query.report_currency,
                )
                exported_rows_count = len(rows) - 1
                filename_stub = (
                    f"giftanalystmarkets_deals_{user.telegram_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                )
                user_id = user.id

            if export_format is ExportFormat.CSV:
                result = ExportFileResult(
                    success=True,
                    message=f"Подготовил CSV-файл. Экспортировано сделок: {exported_rows_count}.",
                    filename=f"{filename_stub}.csv",
                    content=self._build_csv_content(rows),
                    mime_type="text/csv",
                )
            elif export_format is ExportFormat.XLSX:
                result = ExportFileResult(
                    success=True,
                    message=f"Подготовил XLSX-файл. Экспортировано сделок: {exported_rows_count}.",
                    filename=f"{filename_stub}.xlsx",
                    content=self._build_xlsx_content(rows),
                    mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                return ExportFileResult(success=False, message="Неизвестный формат экспорта.")

            if result.success and self._settings is not None:
                await self._log_successful_export(user_id, export_format, exported_rows_count)

            return result

    async def _log_successful_export(
        self,
        user_id: int,
        export_format: ExportFormat,
        rows_exported: int,
    ) -> None:
        """Store a successful export event for daily quota accounting."""

        async with self._session_maker() as session:
            export_log_repo = ExportLogRepository(session)
            async with session.begin():
                await export_log_repo.create(
                    user_id=user_id,
                    export_format=export_format,
                    rows_exported=rows_exported,
                )

    async def _get_today_export_count(
        self,
        export_log_repo: ExportLogRepository,
        user_id: int,
    ) -> int:
        """Return the number of successful exports for the current business day."""

        start_at, end_at, _ = self._get_day_window_utc()
        return await export_log_repo.count_between(user_id, start_at=start_at, end_at=end_at)

    def _build_daily_limit_message(self, current_count: int) -> str:
        """Build a user-facing message for exhausted daily export quota."""

        _, _, timezone_name = self._get_day_window_utc()
        limit = self._settings.daily_export_limit if self._settings is not None else 25
        return (
            "<b>Дневной лимит экспорта исчерпан</b>\n\n"
            f"Сегодня уже использовано <b>{current_count}</b> из <b>{limit}</b> экспортов.\n"
            f"Новый лимит откроется после 00:00 по часовому поясу <b>{timezone_name}</b>."
        )

    def _get_day_window_utc(self) -> tuple[datetime, datetime, str]:
        """Return current business-day window converted to UTC."""

        timezone_name = self._settings.business_timezone if self._settings is not None else "UTC"
        try:
            business_tz = ZoneInfo(timezone_name)
        except ZoneInfoNotFoundError:
            business_tz = timezone.utc
            timezone_name = "UTC"

        now_local = datetime.now(business_tz)
        start_local = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
        end_local = start_local + timedelta(days=1)
        return (
            start_local.astimezone(timezone.utc),
            end_local.astimezone(timezone.utc),
            timezone_name,
        )

    async def _get_ton_rate_for_export(self) -> Decimal | None:
        """Return a fresh TON rate when possible, otherwise latest stored snapshot."""

        if self._ton_client is not None:
            payload = await self._ton_client.get_current_rate()
            if payload.success and payload.rate is not None:
                async with self._session_maker() as session:
                    ton_rate_repo = TonRateRepository(session)
                    async with session.begin():
                        await ton_rate_repo.create(payload.rate, payload.source)
                return payload.rate

        async with self._session_maker() as session:
            ton_rate_repo = TonRateRepository(session)
            latest_rate = await ton_rate_repo.get_latest()
            return latest_rate.rate if latest_rate is not None else None

    def _apply_query_filters(self, deals: list[Any], query: ExportQuery) -> list[Any]:
        """Filter deals based on query fields."""

        filtered = list(deals)

        if query.statuses:
            filtered = [deal for deal in filtered if deal.status.value in query.statuses]

        if query.currencies:
            filtered = [deal for deal in filtered if deal.currency.value in query.currencies]

        if query.profit_filter != "any":
            filtered = [
                deal
                for deal in filtered
                if self._matches_profit_filter(deal.net_profit, query.profit_filter)
            ]

        if query.days is not None:
            cutoff = datetime.now(timezone.utc) - timedelta(days=query.days)
            filtered = [deal for deal in filtered if self._to_utc(deal.created_at) >= cutoff]

        if query.limit is not None:
            filtered = filtered[:query.limit]

        return filtered

    def build_full_export_rows(self, deals: list[Any], ton_rate: Decimal | None) -> list[list[str]]:
        """Build rows for full export without user-selected parameters."""

        return self._build_export_rows(
            deals,
            ton_rate,
            selected_fields=list(DEFAULT_EXPORT_FIELDS),
            report_currency=None,
        )

    def _build_export_rows(
        self,
        deals: list[Any],
        ton_rate: Decimal | None,
        *,
        selected_fields: list[str] | None = None,
        report_currency: str | None = None,
    ) -> list[list[str]]:
        """Build tabular export rows with user-friendly column names and values."""

        fields = selected_fields or DEFAULT_EXPORT_FIELDS
        target_currency = normalize_currency(report_currency)

        if target_currency is None:
            buy_header = "Цена покупки"
            sell_header = "Цена продажи"
            fee_header = "Комиссия"
            profit_header = "Чистая прибыль"
            currency_header = "Валюта"
        else:
            buy_header = f"Цена покупки ({target_currency.value})"
            sell_header = f"Цена продажи ({target_currency.value})"
            fee_header = f"Комиссия ({target_currency.value})"
            profit_header = f"Чистая прибыль ({target_currency.value})"
            currency_header = "Валюта отчета"

        def render_ton_rate(deal: Any) -> Decimal | None:
            return getattr(deal, "ton_usd_rate", None) or ton_rate

        def render_amount(amount: Decimal | None, deal: Any) -> str:
            if target_currency is None:
                return self._format_nullable_decimal(amount)
            converted = self._convert_amount(
                amount,
                source_currency=deal.currency,
                target_currency=target_currency,
                ton_rate=render_ton_rate(deal),
            )
            return self._format_nullable_decimal(converted)

        renderers: dict[str, tuple[str, Any]] = {
            "id": ("ID сделки", lambda deal: deal.external_deal_id),
            "item": ("Предмет", lambda deal: deal.item_name),
            "gift_number": ("Номер", lambda deal: getattr(deal, "gift_number", None) or "—"),
            "market": ("Маркет", lambda deal: getattr(deal, "marketplace", None) or "—"),
            "category": ("Категория", lambda deal: deal.category or "—"),
            "buy": (buy_header, lambda deal: render_amount(deal.buy_price, deal)),
            "sell": (sell_header, lambda deal: render_amount(deal.sell_price, deal)),
            "fee": (fee_header, lambda deal: render_amount(deal.fee, deal)),
            "profit": (profit_header, lambda deal: render_amount(deal.net_profit, deal)),
            "margin": (
                "Маржа (%)",
                lambda deal: self._format_margin_percent(deal.net_profit, deal.buy_price),
            ),
            "currency": (
                currency_header,
                lambda deal: target_currency.value if target_currency is not None else deal.currency.value,
            ),
            "ton_rate": (
                "Курс TON (USD)",
                lambda deal: self._format_nullable_decimal(render_ton_rate(deal)),
            ),
            "status": ("Статус", lambda deal: self._format_status(deal.status.value)),
            "opened_at": ("Дата открытия", lambda deal: self._format_datetime(deal.opened_at)),
            "closed_at": ("Дата закрытия", lambda deal: self._format_datetime(deal.closed_at)),
            "created_at": ("Добавлено в бота", lambda deal: self._format_datetime(deal.created_at)),
            "updated_at": ("Обновлено в боте", lambda deal: self._format_datetime(deal.updated_at)),
            "gift_link": (
                "Ссылка на подарок",
                lambda deal: getattr(deal, "gift_url", None) or "—",
            ),
        }

        headers = [renderers[field_key][0] for field_key in fields]
        rows = [headers]

        for deal in deals:
            rows.append([renderers[field_key][1](deal) for field_key in fields])

        return rows

    def _build_csv_content(self, rows: list[list[str]]) -> bytes:
        """Build CSV bytes in UTF-8 with BOM for spreadsheet compatibility."""

        buffer = StringIO(newline="")
        writer = csv.writer(buffer)
        writer.writerows(rows)
        return buffer.getvalue().encode("utf-8-sig")

    def _build_xlsx_content(self, rows: list[list[str]]) -> bytes:
        """Build a styled XLSX workbook in memory."""

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Сделки"

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for row in rows:
            worksheet.append(row)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            values = [str(cell.value or "") for cell in column_cells]
            max_length = max((len(value) for value in values), default=0)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 36)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _matches_profit_filter(self, net_profit: Decimal | None, profit_filter: str) -> bool:
        """Return whether deal net_profit satisfies requested profit filter."""

        if net_profit is None:
            return False

        if profit_filter == "positive":
            return net_profit > 0
        if profit_filter == "negative":
            return net_profit < 0
        if profit_filter == "zero":
            return net_profit == 0
        if profit_filter == "non_negative":
            return net_profit >= 0
        if profit_filter == "non_positive":
            return net_profit <= 0
        return True

    def _to_utc(self, value: datetime) -> datetime:
        """Normalize datetime to UTC for filtering."""

        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc)

    def _convert_amount(
        self,
        amount: Decimal | int | float | None,
        *,
        source_currency: Currency | str,
        target_currency: Currency,
        ton_rate: Decimal | None,
    ) -> Decimal | None:
        """Convert amount from source currency to target currency via USD pivot."""

        return convert_amount(
            amount,
            source_currency=source_currency,
            target_currency=target_currency,
            ton_usd_rate=ton_rate,
        )

    def _format_decimal(self, value: Decimal | int | float) -> str:
        """Format numeric values with two decimal places."""

        decimal_value = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"{decimal_value}"

    def _format_nullable_decimal(self, value: Decimal | int | float | None) -> str:
        """Format nullable numeric values."""

        if value is None:
            return "—"
        return self._format_decimal(value)

    def _format_datetime(self, value: datetime | None) -> str:
        """Format datetimes in a compact user-friendly form."""

        if value is None:
            return "—"
        return value.strftime("%d.%m.%Y %H:%M")

    def _format_margin_percent(
        self,
        net_profit: Decimal | int | float | None,
        buy_price: Decimal | int | float,
    ) -> str:
        """Calculate margin percentage using net profit and buy price."""

        if net_profit is None:
            return "—"

        buy_decimal = Decimal(str(buy_price))
        if buy_decimal == 0:
            return "—"

        margin = (Decimal(str(net_profit)) / buy_decimal) * Decimal("100")
        margin = margin.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return f"{margin}%"

    def _format_status(self, value: str) -> str:
        """Map internal statuses to human-friendly labels."""

        mapping = {
            "open": "Открыта",
            "closed": "Закрыта",
            "cancelled": "Отменена",
        }
        return mapping.get(value, value)
